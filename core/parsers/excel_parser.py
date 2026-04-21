from __future__ import annotations

import io
from openpyxl import load_workbook


def _sheet_to_markdown(ws) -> str:
    """Convert a worksheet to a Markdown table string."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    def cell_str(v) -> str:
        if v is None:
            return ""
        return str(v).replace("|", "\\|")

    header = rows[0]
    lines = [
        "| " + " | ".join(cell_str(c) for c in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in rows[1:]:
        padded = list(row) + [None] * (len(header) - len(row))
        lines.append("| " + " | ".join(cell_str(c) for c in padded[:len(header)]) + " |")

    return "\n".join(lines)


def parse_excel_by_sheet(file_bytes: bytes) -> list[dict]:
    """Parse Excel file, returning one chunk per sheet in Markdown table format."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    chunks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        content = _sheet_to_markdown(ws)
        if content:
            chunks.append({
                "content": content,
                "metadata": {"sheet": sheet_name},
            })
    wb.close()
    return chunks


def parse_excel_by_sheet_with_row_split(
    file_bytes: bytes, max_rows: int = 100
) -> list[dict]:
    """Parse Excel file, splitting large sheets into multiple chunks by row count."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = rows[0]

        def cell_str(v) -> str:
            if v is None:
                return ""
            return str(v).replace("|", "\\|")

        header_line = "| " + " | ".join(cell_str(c) for c in header) + " |"
        separator = "| " + " | ".join("---" for _ in header) + " |"

        data_rows = rows[1:]
        part_idx = 0
        for start in range(0, max(len(data_rows), 1), max_rows):
            batch = data_rows[start:start + max_rows]
            if not batch:
                continue
            lines = [header_line, separator]
            for row in batch:
                padded = list(row) + [None] * (len(header) - len(row))
                lines.append("| " + " | ".join(cell_str(c) for c in padded[:len(header)]) + " |")
            chunks.append({
                "content": "\n".join(lines),
                "metadata": {
                    "sheet": sheet_name,
                    "part": part_idx,
                    "rows": f"{start + 2}-{start + 1 + len(batch)}",
                },
            })
            part_idx += 1

    wb.close()
    return chunks
