from __future__ import annotations

import io

import fitz

PDF_MIME = "application/pdf"
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLS_MIME = "application/vnd.ms-excel"


def _resolve_bounds(total: int, start_1b: int | None, end_1b: int | None) -> tuple[int, int]:
    """Return inclusive (start_0b, end_0b). total is total count of pages/sheets."""
    if total <= 0:
        raise ValueError("Document has 0 pages/sheets.")
    start = (start_1b - 1) if start_1b else 0
    end = (end_1b - 1) if end_1b else (total - 1)
    start = max(0, start)
    end = min(total - 1, end)
    if start > end:
        raise ValueError(f"Invalid page range: start={start_1b}, end={end_1b}, total={total}.")
    return start, end


def extract_pdf_file(file_bytes: bytes, start_1b: int | None, end_1b: int | None) -> tuple[bytes, str]:
    """Extract a contiguous page range from a PDF and return new PDF bytes + mime."""
    src = fitz.open(stream=file_bytes, filetype="pdf")
    total = src.page_count
    start, end = _resolve_bounds(total, start_1b, end_1b)

    dst = fitz.open()
    dst.insert_pdf(src, from_page=start, to_page=end)
    out = dst.tobytes(garbage=3, deflate=True)
    dst.close()
    src.close()
    return out, PDF_MIME


def extract_excel_file(file_bytes: bytes, start_1b: int | None, end_1b: int | None) -> tuple[bytes, str]:
    """Keep sheets in 1-based index range [start, end], remove the rest. Returns xlsx bytes."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes))
    total = len(wb.sheetnames)
    start, end = _resolve_bounds(total, start_1b, end_1b)

    keep = set(wb.sheetnames[start:end + 1])
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]

    if not wb.sheetnames:
        raise ValueError("No sheets remain after filtering.")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), XLSX_MIME


def _count_word_pages(doc) -> int:
    """Count pages as (manual page_break count + 1)."""
    from docx.oxml.ns import qn
    count = 1
    for br in doc.element.body.iter(qn("w:br")):
        if br.get(qn("w:type")) == "page":
            count += 1
    return count


def _element_page_index(body, element_to_page: dict) -> dict:
    """Walk body children and assign each top-level element (p/tbl) a page index (0-based)."""
    from docx.oxml.ns import qn
    current_page = 0
    for child in list(body):
        element_to_page[id(child)] = current_page
        for br in child.iter(qn("w:br")):
            if br.get(qn("w:type")) == "page":
                current_page += 1
    return element_to_page


def extract_word_file(file_bytes: bytes, start_1b: int | None, end_1b: int | None) -> tuple[bytes, str]:
    """
    Remove paragraphs/tables outside the selected page range.
    Page boundaries are determined by manual page breaks (<w:br w:type='page'>).
    Section properties (sectPr) and headers/footers are preserved.
    """
    from docx import Document
    from docx.oxml.ns import qn

    doc = Document(io.BytesIO(file_bytes))
    body = doc.element.body

    total = _count_word_pages(doc)
    start, end = _resolve_bounds(total, start_1b, end_1b)

    element_to_page: dict = {}
    _element_page_index(body, element_to_page)

    removable_tags = {qn("w:p"), qn("w:tbl")}
    for child in list(body):
        if child.tag not in removable_tags:
            continue
        page = element_to_page.get(id(child), 0)
        if page < start or page > end:
            body.remove(child)

    out = io.BytesIO()
    doc.save(out)
    return out.getvalue(), DOCX_MIME


def extract_file(file_bytes: bytes, file_ext: str, start_1b: int | None, end_1b: int | None) -> tuple[bytes, str]:
    ext = file_ext.lower().lstrip(".")
    if ext == "pdf":
        return extract_pdf_file(file_bytes, start_1b, end_1b)
    if ext == "docx":
        return extract_word_file(file_bytes, start_1b, end_1b)
    if ext in ("xlsx", "xls"):
        return extract_excel_file(file_bytes, start_1b, end_1b)
    raise ValueError(f"Unsupported file type for page_file mode: .{ext}")
